"""
Amazon Seller Niche + Contact Pipeline (2 stages, streamed writes)

Input: CSV/XLSX with headers:
Seller ID, Seller Link, PL URL, Seller Name, Website, Estimated Monthly Revenue,
First Name, Last Name, Title, Email, Source, Phone Number, Country, State

Stage 1 output: input headers + niche rank
Stage 2 output: input headers + niche rank + text/apollo contact columns

Important:
- Product id extraction is done from PL URL only.
- Missing fields never remove a row; stage fields are set to null.
"""

import csv
import html
import os
import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook

from config_env import override_from_env

CONFIG = {
    "INPUT_FILE": "data/seller_niche_input.csv",
    "STAGE1_OUTPUT_FILE": "data/seller_niche_stage1.csv",
    "STAGE2_OUTPUT_FILE": "data/seller_niche_stage2.csv",
    "AMAZON_DOMAIN": "amazon.de",
    "API_TIMEOUT": 60,
    "WAIT_BETWEEN_REQUESTS": 1,
    "WAIT_BETWEEN_BATCHES": 1,
    "HTTP": "true",
    "DEVICE": "desktop",
    "SCROLL_TO_BOTTOM": "true",
    "WAIT_BOTTOM_CAROUSEL": "true",
    "WAIT_FOR_OFFERS": "true",
    "WAIT_FOR_VIDEO": "true",
    "MODEL": "gpt-5-mini",
    "REASONING_EFFORT": "medium",
    "MAX_OUTPUT_TOKENS": 4000,
    "APOLLO_BASE_URL": "https://api.apollo.io/api/v1",
    "APOLLO_API_TIMEOUT": 60,
    "APOLLO_HTTP_RETRIES": 3,
    "APOLLO_SEARCH_BY_NAME_FALLBACK": False,
    "STAGE1_WRITE_BATCH_SIZE": 100,
    "STAGE2_WRITE_BATCH_SIZE": 50,
    # When false, only Stage 1 runs (niche rank). No GPT text extract, Apollo, or APOLLO_API_KEY required.
    "RUN_STAGE2_PERSON_STAGE": True,
}

INPUT_HEADERS = [
    "Seller ID",
    "Seller Link",
    "PL URL",
    "Seller Name",
    "Website",
    "Estimated Monthly Revenue",
    "First Name",
    "Last Name",
    "Title",
    "Email",
    "Source",
    "Phone Number",
    "Country",
    "State",
]

STAGE1_HEADERS = INPUT_HEADERS + ["niche rank"]

STAGE2_EXTRA_HEADERS = [
    "niche rank",
    "text email",
    "text number",
    "text name",
    "text title",
    "apollo name",
    "apollo title",
    "apollo email",
]
STAGE2_HEADERS = INPUT_HEADERS + STAGE2_EXTRA_HEADERS

BASE_URL = "https://ecom.webscrapingapi.com/v1"
API_KEY = ""
APOLLO_API_KEY = ""
PUBLIC_EMAIL_DOMAINS = frozenset()
client = None


def config_run_stage2_person_stage() -> bool:
    """
    True = run Stage 2 (GPT text + Apollo). Must handle real bool and string env values:
    bool('false') is True in Python, so never use bool() on raw config.
    """
    v = CONFIG.get("RUN_STAGE2_PERSON_STAGE", True)
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("0", "false", "no", "off", "none", ""):
        return False
    if s in ("1", "true", "yes", "on"):
        return True
    return True


def null_str(v) -> str:
    if v is None:
        return "null"
    s = str(v).strip()
    return s if s else "null"


def text_field(v) -> str:
    if v is None:
        return "null"
    s = html.unescape(str(v)).strip()
    return s if s else "null"


def load_public_email_domains():
    global PUBLIC_EMAIL_DOMAINS
    p = Path(__file__).resolve().parent / "public_email_domains.txt"
    vals = set()
    if p.is_file():
        for line in p.read_text(encoding="utf-8").splitlines():
            s = line.strip().lower()
            if s and not s.startswith("#"):
                vals.add(s)
    vals.update({"gmail.com", "googlemail.com", "yahoo.com", "hotmail.com", "outlook.com", "gmx.de", "web.de"})
    PUBLIC_EMAIL_DOMAINS = frozenset(vals)


def extract_company_domain_from_email(email: str) -> Optional[str]:
    s = (email or "").strip().strip("'").strip('"')
    if "@" not in s:
        return None
    d = s.split("@", 1)[1].strip().lower().split("/")[0].split("?")[0].rstrip(">")
    if not d or d in PUBLIC_EMAIL_DOMAINS:
        return None
    return d


def extract_product_id_from_link(link: str) -> str:
    if not link:
        return ""
    s = link.strip()
    m = re.search(r"/dp/([A-Z0-9]{10})(?:[/?]|$)", s, flags=re.IGNORECASE)
    if m:
        return m.group(1).upper()
    m = re.search(r"(?:asin|product_id)=([A-Z0-9]{10})", s, flags=re.IGNORECASE)
    if m:
        return m.group(1).upper()
    return ""


def fetch_json(url: str):
    try:
        r = requests.get(url, timeout=int(CONFIG["API_TIMEOUT"]))
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"    API error: {e}")
        return None


def build_product_url(product_id: str) -> str:
    params = {
        "engine": "amazon",
        "api_key": API_KEY,
        "type": "product",
        "product_id": product_id,
        "amazon_domain": CONFIG["AMAZON_DOMAIN"],
        "scroll_to_bottom": CONFIG["SCROLL_TO_BOTTOM"],
        "wait_bottom_carousel": CONFIG["WAIT_BOTTOM_CAROUSEL"],
        "wait_for_offers": CONFIG["WAIT_FOR_OFFERS"],
        "wait_for_video": CONFIG["WAIT_FOR_VIDEO"],
        "http": CONFIG["HTTP"],
        "device": CONFIG["DEVICE"],
    }
    return BASE_URL + "?" + "&".join([f"{k}={v}" for k, v in params.items()])


def build_seller_url(seller_id: str) -> str:
    params = {
        "engine": "amazon",
        "api_key": API_KEY,
        "type": "seller_profile",
        "seller_id": seller_id,
        "amazon_domain": CONFIG["AMAZON_DOMAIN"],
        "http": CONFIG["HTTP"],
        "device": CONFIG["DEVICE"],
    }
    return BASE_URL + "?" + "&".join([f"{k}={v}" for k, v in params.items()])


SYSTEM_PROMPT_NICHE = """Classify Amazon product niche quality from title, rating and review count.

Return one word only: good, neutral, or bad.

Rules:
- good: strong proven demand and healthy rating signal (e.g., high review volume with solid rating).
- neutral: mixed/uncertain signal (e.g., very few reviews even with high rating).
- bad: weak signal (low rating and/or poor traction signal).
- If both rating and review count are missing or unusable, output bad.
- Use only the provided fields; no extra assumptions.
"""


def _niche_has_usable_rating(rating: str) -> bool:
    s = (rating or "").strip().lower().strip("'").strip('"')
    if not s or s in ("null", "none", "n/a", "na", "-"):
        return False
    try:
        return float(s.replace(",", ".")) > 0
    except (ValueError, TypeError):
        m = re.search(r"(\d+[.,]\d+|\d+)", s)
        return bool(m)


def _niche_has_usable_review_count(review_count: str) -> bool:
    s = (review_count or "").strip().lower().strip("'").strip('"')
    if not s or s in ("null", "none", "n/a", "na", "-"):
        return False
    try:
        return int(float(s.replace(",", "."))) > 0
    except (ValueError, TypeError):
        return False


def niche_rank_without_signal(rating: str, review_count: str) -> bool:
    """True when there is no usable rating and no usable review count (force bad, skip GPT)."""
    return not _niche_has_usable_rating(rating) and not _niche_has_usable_review_count(review_count)


def call_gpt_niche_rank(title: str, rating: str, review_count: str) -> str:
    if niche_rank_without_signal(rating, review_count):
        return "bad"
    user_prompt = (
        f"Title: {title or 'null'}\n"
        f"Rating: {rating or 'null'}\n"
        f"Review count: {review_count or 'null'}\n"
        "Output exactly one token: good | neutral | bad"
    )
    try:
        response = client.responses.create(
            model=CONFIG["MODEL"],
            reasoning={"effort": CONFIG["REASONING_EFFORT"]},
            input=[
                {"role": "system", "content": SYSTEM_PROMPT_NICHE},
                {"role": "user", "content": user_prompt},
            ],
            max_output_tokens=int(CONFIG["MAX_OUTPUT_TOKENS"]),
        )
        out = (response.output_text or "").strip().lower()
        if out in {"good", "neutral", "bad"}:
            return out
        if "good" in out:
            return "good"
        if "bad" in out:
            return "bad"
        return "neutral"
    except Exception as e:
        print(f"    GPT niche error: {e}")
        return "neutral"


SYSTEM_PROMPT_TEXT = """Extract contact details from seller text only.

Input has seller about + seller description.
Return exactly one CSV data line:
text email;text number;text name;text title

Rules:
- Use only given texts.
- If missing, output null.
- Return only ONE person (best relevant person).
- If person is unclear, use null for name and title.
- No extra text.
"""


def call_gpt_text_extract(seller_about: str, seller_description: str) -> Tuple[str, str, str, str]:
    user_prompt = (
        f"seller about: {seller_about or 'null'}\n"
        f"seller description: {seller_description or 'null'}\n"
        "Output: text email;text number;text name;text title"
    )
    try:
        response = client.responses.create(
            model=CONFIG["MODEL"],
            reasoning={"effort": CONFIG["REASONING_EFFORT"]},
            input=[
                {"role": "system", "content": SYSTEM_PROMPT_TEXT},
                {"role": "user", "content": user_prompt},
            ],
            max_output_tokens=int(CONFIG["MAX_OUTPUT_TOKENS"]),
        )
        raw = (response.output_text or "").strip().strip("`")
        line = raw.splitlines()[0] if raw else ""
        parts = [p.strip() for p in line.split(";")]
        if len(parts) >= 4:
            return tuple(null_str(p) for p in parts[:4])
    except Exception as e:
        print(f"    GPT text extraction error: {e}")
    return "null", "null", "null", "null"


SYSTEM_PROMPT_APOLLO_PICK = """Choose the best contact for Amazon/e-commerce creative outreach.

Input rows are: id | title
Return one line only:
<apollo_person_id>, <title>
or null if none is suitable.
"""


def apollo_post(path: str, json_body=None):
    url = CONFIG["APOLLO_BASE_URL"].rstrip("/") + path
    headers = {"x-api-key": APOLLO_API_KEY, "Content-Type": "application/json", "Cache-Control": "no-cache"}
    last_err = None
    for attempt in range(1, int(CONFIG["APOLLO_HTTP_RETRIES"]) + 1):
        try:
            r = requests.post(url, headers=headers, json=json_body or {}, timeout=int(CONFIG["APOLLO_API_TIMEOUT"]))
            r.raise_for_status()
            return r.json()
        except Exception as e:
            last_err = e
            print(f"    Apollo attempt {attempt}/{CONFIG['APOLLO_HTTP_RETRIES']}: {e}")
            time.sleep(float(CONFIG["WAIT_BETWEEN_REQUESTS"]))
    print(f"    Apollo failed: {last_err}")
    return None


def apollo_search_company(domain: Optional[str], company_name: str):
    if domain:
        data = apollo_post("/mixed_companies/search", {"q_organization_domains_list": [domain]})
    elif CONFIG.get("APOLLO_SEARCH_BY_NAME_FALLBACK", False) and company_name:
        data = apollo_post("/mixed_companies/search", {"q_organization_name": company_name})
    else:
        return None
    if not data:
        return None
    orgs = data.get("organizations") or []
    if orgs:
        return orgs[0]
    accounts = data.get("accounts") or []
    if accounts and isinstance(accounts[0], dict):
        oid = (accounts[0].get("organization_id") or "").strip()
        if oid:
            return {"id": oid}
    return None


def apollo_list_people(org_id: str) -> List[dict]:
    data = apollo_post("/mixed_people/api_search", {"organization_ids": [org_id], "per_page": 100})
    return (data or {}).get("people") or []


def apollo_match_person(person_id: str):
    data = apollo_post("/people/match", {"id": person_id})
    if not data:
        return None
    return data.get("person") or data


def call_gpt_pick_apollo(people: List[Tuple[str, str]]) -> Optional[str]:
    lines = ["id | title"]
    lines.extend([f"{pid} | {title}" for pid, title in people])
    lines.append("Output one line: id, title OR null")
    try:
        response = client.responses.create(
            model=CONFIG["MODEL"],
            reasoning={"effort": CONFIG["REASONING_EFFORT"]},
            input=[
                {"role": "system", "content": SYSTEM_PROMPT_APOLLO_PICK},
                {"role": "user", "content": "\n".join(lines)},
            ],
            max_output_tokens=int(CONFIG["MAX_OUTPUT_TOKENS"]),
        )
        out = (response.output_text or "").strip()
        if not out or out.lower() == "null":
            return None
        first = out.splitlines()[0]
        pid = first.split(",", 1)[0].strip()
        return pid or None
    except Exception as e:
        print(f"    GPT Apollo pick error: {e}")
        return None


def iter_input_rows(path: str):
    p = Path(path)
    if p.suffix.lower() == ".xlsx":
        wb = load_workbook(filename=p, data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = {}
            for i, h in enumerate(header):
                if h:
                    rec[h] = "" if i >= len(row) or row[i] is None else str(row[i]).strip()
            yield rec
        return
    with open(p, "r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            yield dict(row)


def flush_csv_rows(path: str, headers: List[str], rows: List[Dict[str, str]], mode: str):
    if not rows:
        return
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, mode, encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        if mode == "w":
            w.writeheader()
        w.writerows(rows)


def validate_headers_from_row(first_row: Dict[str, str]):
    if not first_row:
        raise ValueError("Input file has no data rows.")
    missing = [h for h in INPUT_HEADERS if h not in first_row]
    if missing:
        raise ValueError(f"Missing required headers: {missing}")


def stage1_build_rank_file():
    print("\n" + "=" * 90)
    print("STAGE 1 - Build rank file (streamed)")
    print("=" * 90)
    print(f"Input:  {CONFIG['INPUT_FILE']}")
    print(f"Output: {CONFIG['STAGE1_OUTPUT_FILE']}")

    row_iter = iter_input_rows(CONFIG["INPUT_FILE"])
    first = next(row_iter, None)
    validate_headers_from_row(first)

    mode = "w"
    batch_size = int(CONFIG["STAGE1_WRITE_BATCH_SIZE"])
    buf = []
    written = 0

    def process_row(row: Dict[str, str], i: int):
        seller_id = (row.get("Seller ID") or "").strip()
        pl_url = (row.get("PL URL") or "").strip()
        product_id = extract_product_id_from_link(pl_url)
        title = "null"
        rating = "null"
        review_count = "null"
        niche_rank = "bad"
        print(f"[S1 row {i}] seller_id={seller_id or 'null'} product_id={product_id or 'null'}")

        if product_id:
            pdata = fetch_json(build_product_url(product_id))
            product = (pdata or {}).get("product_results", {})
            if isinstance(product, dict):
                title = text_field(product.get("title"))
                rating = null_str(product.get("rating"))
                review_count = null_str(product.get("ratings_total"))
            if niche_rank_without_signal(rating, review_count):
                niche_rank = "bad"
            else:
                niche_rank = call_gpt_niche_rank(title, rating, review_count)
                time.sleep(float(CONFIG["WAIT_BETWEEN_BATCHES"]))

        out = {h: null_str(row.get(h, "null")) for h in INPUT_HEADERS}
        out["niche rank"] = niche_rank
        time.sleep(float(CONFIG["WAIT_BETWEEN_REQUESTS"]))
        return out

    i = 1
    if first is not None:
        buf.append(process_row(first, i))
    for row in row_iter:
        i += 1
        buf.append(process_row(row, i))
        if len(buf) >= batch_size:
            flush_csv_rows(CONFIG["STAGE1_OUTPUT_FILE"], STAGE1_HEADERS, buf, mode)
            written += len(buf)
            print(f"   Flushed {len(buf)} rows (total: {written})")
            buf = []
            mode = "a"

    if buf:
        flush_csv_rows(CONFIG["STAGE1_OUTPUT_FILE"], STAGE1_HEADERS, buf, mode)
        written += len(buf)
        print(f"   Final flush {len(buf)} rows (total: {written})")

    print(f"Stage 1 complete. Rows written: {written}")


def stage2_add_person_data():
    print("\n" + "=" * 90)
    print("STAGE 2 - Add person data (GPT + Apollo, streamed)")
    print("=" * 90)
    print(f"Input:  {CONFIG['STAGE1_OUTPUT_FILE']}")
    print(f"Output: {CONFIG['STAGE2_OUTPUT_FILE']}")

    mode = "w"
    batch_size = int(CONFIG["STAGE2_WRITE_BATCH_SIZE"])
    buf = []
    written = 0

    with open(CONFIG["STAGE1_OUTPUT_FILE"], "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, 1):
            seller_id = (row.get("Seller ID") or "").strip()
            seller_about = "null"
            seller_description = "null"
            text_email = "null"
            text_number = "null"
            text_name = "null"
            text_title = "null"
            apollo_name = "null"
            apollo_title = "null"
            apollo_email = "null"

            print(f"[S2 row {i}] seller_id={seller_id or 'null'}")

            if seller_id and seller_id.lower() != "null":
                sdata = fetch_json(build_seller_url(seller_id))
                details = (sdata or {}).get("seller_profile", {}).get("seller_details", {})
                if isinstance(details, dict):
                    seller_description = text_field(details.get("detailed_information"))
                    seller_about = text_field(details.get("about_this_seller"))
                text_email, text_number, text_name, text_title = call_gpt_text_extract(seller_about, seller_description)
                time.sleep(float(CONFIG["WAIT_BETWEEN_BATCHES"]))

                domain = extract_company_domain_from_email(text_email)
                company_name = (row.get("Seller Name") or "").strip()
                org = apollo_search_company(domain, company_name)
                time.sleep(float(CONFIG["WAIT_BETWEEN_REQUESTS"]))
                if org and isinstance(org, dict) and org.get("id"):
                    people = apollo_list_people(org["id"])
                    people_lines = []
                    for p in people:
                        if not isinstance(p, dict):
                            continue
                        pid = (p.get("id") or "").strip()
                        ptitle = (p.get("title") or "").strip()
                        if pid and ptitle:
                            people_lines.append((pid, ptitle))
                    if people_lines:
                        best_pid = call_gpt_pick_apollo(people_lines)
                        if best_pid:
                            person = apollo_match_person(best_pid)
                            if isinstance(person, dict):
                                email = (person.get("email") or "").strip()
                                status = (person.get("email_status") or "").strip().lower()
                                if email and status == "verified":
                                    apollo_name = (person.get("name") or "").strip() or (
                                        ((person.get("first_name") or "").strip() + " " + (person.get("last_name") or "").strip()).strip()
                                    )
                                    apollo_title = (person.get("title") or "").strip() or "null"
                                    apollo_email = email

            out = {h: null_str(row.get(h, "null")) for h in INPUT_HEADERS}
            out["niche rank"] = null_str(row.get("niche rank"))
            out["text email"] = null_str(text_email)
            out["text number"] = null_str(text_number)
            out["text name"] = null_str(text_name)
            out["text title"] = null_str(text_title)
            out["apollo name"] = null_str(apollo_name)
            out["apollo title"] = null_str(apollo_title)
            out["apollo email"] = null_str(apollo_email)
            buf.append(out)
            time.sleep(float(CONFIG["WAIT_BETWEEN_REQUESTS"]))

            if len(buf) >= batch_size:
                flush_csv_rows(CONFIG["STAGE2_OUTPUT_FILE"], STAGE2_HEADERS, buf, mode)
                written += len(buf)
                print(f"   Flushed {len(buf)} rows (total: {written})")
                buf = []
                mode = "a"

    if buf:
        flush_csv_rows(CONFIG["STAGE2_OUTPUT_FILE"], STAGE2_HEADERS, buf, mode)
        written += len(buf)
        print(f"   Final flush {len(buf)} rows (total: {written})")

    print(f"Stage 2 complete. Rows written: {written}")


def process():
    stage1_build_rank_file()
    if config_run_stage2_person_stage():
        stage2_add_person_data()
    else:
        print("\n" + "=" * 90)
        print("STAGE 2 skipped (RUN_STAGE2_PERSON_STAGE=false)")
        print("=" * 90)


if __name__ == "__main__":
    load_dotenv(Path(__file__).resolve().parent / ".env")
    # Fixed env prefix for this script
    override_from_env(CONFIG, env_prefix="SELLERNICHE_")
    CONFIG["RUN_STAGE2_PERSON_STAGE"] = config_run_stage2_person_stage()
    print(f"SELLERNICHE RUN_STAGE2_PERSON_STAGE → {CONFIG['RUN_STAGE2_PERSON_STAGE']}")

    API_KEY = (os.getenv("API_KEY") or "").strip()
    OPENAI_API_KEY = (os.getenv("OPENAI_API_KEY") or "").strip()
    APOLLO_API_KEY = (os.getenv("APOLLO_API_KEY") or "").strip()
    if not API_KEY:
        raise ValueError("API_KEY not found in .env")
    if not OPENAI_API_KEY:
        raise ValueError("OPENAI_API_KEY not found in .env")
    run_stage2 = CONFIG["RUN_STAGE2_PERSON_STAGE"]
    if run_stage2 and not APOLLO_API_KEY:
        raise ValueError("APOLLO_API_KEY not found in .env (required when RUN_STAGE2_PERSON_STAGE is true)")

    if run_stage2:
        load_public_email_domains()
    client = OpenAI(api_key=OPENAI_API_KEY)
    # Stage outputs are CSV-only for streamed realtime writes.
    if not str(CONFIG["STAGE1_OUTPUT_FILE"]).lower().endswith(".csv"):
        raise ValueError("STAGE1_OUTPUT_FILE must be .csv")
    if run_stage2 and not str(CONFIG["STAGE2_OUTPUT_FILE"]).lower().endswith(".csv"):
        raise ValueError("STAGE2_OUTPUT_FILE must be .csv")
    process()
